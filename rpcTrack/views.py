from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages 
from django.db import IntegrityError
from django.core.exceptions import ObjectDoesNotExist
from django.utils.datastructures import MultiValueDictKeyError
from django.views.generic.edit import FormView

### modules ###
import platform
import os
import subprocess
import openpyxl
from openpyxl.styles import Font

### imports ###
from .models import UE_LIST, BTS_PC, TM500_PC, BTS_INFO, BTS_MODULES, TM500_INFO

# Create your views here.

def index(request):
    # display all remote pcs #
    getBTSPC = BTS_PC.objects.all()  # get bts pcs
    getUE = UE_LIST.objects.all()    # get ue types
    getTMPC = TM500_PC.objects.all() # get tm500 pcs
    getBTSINFO = BTS_INFO.objects.all() # get bts pc info
    getBTSMOD = BTS_MODULES.objects.all()
    getTMINFO = TM500_INFO.objects.all()
    # newnew = btsPCInfo(serverName=btsPC.objects.get(serverName='10.12.25.11'), WCDMAPilot = '12345')
    # newnew.save()

    if request.method == "POST":
        operation = request.POST['operation']

        ##############################################################
        ################### BTS PC OPERATIONS ########################
        ##############################################################
        if operation == "add-bts":
            server_name = request.POST['server_name']
            display_name = request.POST['display_name']
            ussername = request.POST['username']
            pass_word = request.POST['password']
            do_main = request.POST['do_main']
            u_e = request.POST['u_e']
            sport = request.POST['sport']
            txta = request.POST['txta']
            rack = request.POST['rack']
            s_n = request.POST['s_n']
            p_n = request.POST['p_n']
            a_n = request.POST['a_n']
            # tm_500 = request.POST['tm_500']
            # t_t = request.POST['t_t']

            print(server_name, display_name, ussername, pass_word, do_main, u_e, sport, txta, rack)
            try:
                '''newBTS = BTS_PC(ip=server_name, display_name=display_name, username=ussername, password=pass_word, domain=do_main, ue_type=u_e, switch_port=sport, notes=txta, rack=rack, serial_number=s_n, product_number=p_n, asset_number=a_n)
                newBTS.save()
                print("added")'''
                if u_e == 'TM500':
                    tm_500 = request.POST['tm_500']
                    print(tm_500)
                    #newBTS = BTS_PC(ip=server_name, display_name=display_name, username=ussername, password=pass_word, domain=do_main, ue_type=u_e, switch_port=sport, notes=txta, rack=rack, serial_number=s_n, product_number=p_n, asset_number=a_n, tm500_pc_id=tm500PC.objects.get(serverName=tm_500))
                    #newBTS.save()
                    if BTS_PC.objects.filter(ip = server_name).exists():
                        messages.error(request, 'Cannot Add BTS PC')
                    else:
                        try:
                            bts_info = request.POST['bts_info_id']
                            print("bts_info: ", bts_info)   
                            newBTS = BTS_PC(ip=server_name, display_name=display_name, username=ussername, password=pass_word, domain=do_main, ue_type=u_e, switch_port=sport, notes=txta, rack=rack, serial_number=s_n, product_number=p_n, asset_number=a_n, tm500_pc_id=TM500_PC.objects.get(ip=tm_500), bts_info_id=BTS_INFO.objects.get(bts_name=bts_info))
                            newBTS.save()
                            count = 1;

                            for mm in getBTSMOD:
                                if str(mm.bts_info_id.bts_name) == str(bts_info):
                                    if count == 1:
                                        #bts_info_id=BTS_INFO.objects.get(bts_name=bts_info)
                                        BTS_PC.objects.filter(ip=server_name).update(bts_mod1=BTS_MODULES.objects.get(i_d=mm.i_d))
                                    elif count == 2:
                                        BTS_PC.objects.filter(ip=server_name).update(bts_mod2=BTS_MODULES.objects.get(i_d=mm.i_d))
                                    elif count == 3:
                                        BTS_PC.objects.filter(ip=server_name).update(bts_mod3=BTS_MODULES.objects.get(i_d=mm.i_d))
                                    count = count + 1
                        except:
                            print("no bts info")
                            newBTS = BTS_PC(ip=server_name, display_name=display_name, username=ussername, password=pass_word, domain=do_main, ue_type=u_e, switch_port=sport, notes=txta, rack=rack, serial_number=s_n, product_number=p_n, asset_number=a_n, tm500_pc_id=TM500_PC.objects.get(ip=tm_500))
                            newBTS.save()
                        messages.success(request, 'Successfully Added')
                elif u_e == "Test Terminal":
                    t_t = request.POST['t_t']
                    print(t_t)
                    if BTS_PC.objects.filter(ip = server_name).exists():
                        messages.error(request, 'Cannot Add BTS PC')
                    else:
                        try:
                            bts_info = request.POST['bts_info_id']
                            print("bts_info: ", bts_info)
                            newBTS = BTS_PC(ip=server_name, display_name=display_name, username=ussername, password=pass_word, domain=do_main, ue_type=u_e, switch_port=sport, notes=txta, rack=rack, serial_number=s_n, product_number=p_n, asset_number=a_n, tt_info_id=BTS_INFO.objects.get(bts_name=t_t), bts_info_id=BTS_INFO.objects.get(bts_name=bts_info))
                            newBTS.save()

                            countbts = 1;
                            counttt = 1;

                            for mm in getBTSMOD:
                                if str(mm.bts_info_id.bts_name) == str(bts_info):
                                    if countbts == 1:
                                        #bts_info_id=BTS_INFO.objects.get(bts_name=bts_info)
                                        BTS_PC.objects.filter(ip=server_name).update(bts_mod1=BTS_MODULES.objects.get(i_d=mm.i_d))
                                    elif countbts == 2:
                                        BTS_PC.objects.filter(ip=server_name).update(bts_mod2=BTS_MODULES.objects.get(i_d=mm.i_d))
                                    elif countbts == 3:
                                        BTS_PC.objects.filter(ip=server_name).update(bts_mod3=BTS_MODULES.objects.get(i_d=mm.i_d))
                                    countbts = countbts + 1
                                if str(mm.bts_info_id.bts_name) == str(t_t):
                                    if counttt == 1:
                                        #bts_info_id=BTS_INFO.objects.get(bts_name=bts_info)
                                        BTS_PC.objects.filter(ip=server_name).update(tt_mod1=BTS_MODULES.objects.get(i_d=mm.i_d))
                                    elif counttt == 2:
                                        BTS_PC.objects.filter(ip=server_name).update(tt_mod2=BTS_MODULES.objects.get(i_d=mm.i_d))
                                    elif counttt == 3:
                                        BTS_PC.objects.filter(ip=server_name).update(tt_mod3=BTS_MODULES.objects.get(i_d=mm.i_d))
                                    counttt = counttt + 1
                        except:
                            print("no bts info")
                            newBTS = BTS_PC(ip=server_name, display_name=display_name, username=ussername, password=pass_word, domain=do_main, ue_type=u_e, switch_port=sport, notes=txta, rack=rack, serial_number=s_n, product_number=p_n, asset_number=a_n, tt_info_id=BTS_INFO.objects.get(bts_name=t_t))
                            newBTS.save()
                            counttt = 1;

                            for mm in getBTSMOD:
                                if str(mm.bts_info_id.bts_name) == str(t_t):
                                    if counttt == 1:
                                        #bts_info_id=BTS_INFO.objects.get(bts_name=bts_info)
                                        BTS_PC.objects.filter(ip=server_name).update(tt_mod1=BTS_MODULES.objects.get(i_d=mm.i_d))
                                    elif counttt == 2:
                                        BTS_PC.objects.filter(ip=server_name).update(tt_mod2=BTS_MODULES.objects.get(i_d=mm.i_d))
                                    elif counttt == 3:
                                        BTS_PC.objects.filter(ip=server_name).update(tt_mod3=BTS_MODULES.objects.get(i_d=mm.i_d))
                                    counttt = counttt + 1
                        messages.success(request, 'Successfully Added')
            except:
                    messages.error(request, 'ERROR')

        elif operation == "update-bts":
            server_name = request.POST['editID']
            display_name = request.POST['editDN']
            ussername = request.POST['editUN']
            pass_word = request.POST['editPW']
            do_main = request.POST['editDM']
            u_e = request.POST['editUE']
            sport = request.POST['editSPort']
            txta = request.POST['editNotes']
            rack = request.POST['editRack']
            s_n = request.POST['editSNum']
            p_n = request.POST['editPNum']
            a_n = request.POST['editANum']

            '''
            count = 1;

                            for mm in getBTSMOD:
                                if str(mm.bts_info_id.bts_name) == str(bts_info):
                                    if count == 1:
                                        #bts_info_id=BTS_INFO.objects.get(bts_name=bts_info)
                                        BTS_PC.objects.filter(ip=server_name).update(bts_mod1=BTS_MODULES.objects.get(i_d=mm.i_d))
                                    elif count == 2:
                                        BTS_PC.objects.filter(ip=server_name).update(bts_mod2=BTS_MODULES.objects.get(i_d=mm.i_d))
                                    elif count == 3:
                                        BTS_PC.objects.filter(ip=server_name).update(bts_mod3=BTS_MODULES.objects.get(i_d=mm.i_d))
                                    count = count + 1'''

            try:
                bts_info = request.POST['editBTS']
                print(bts_info)
                try:
                    if u_e == 'TM500':
                        tm_500 = request.POST['editTM500']
                        BTS_PC.objects.filter(ip=server_name).update(display_name=display_name, username=ussername, password=pass_word, domain=do_main, ue_type=u_e, switch_port=sport, notes=txta, rack=rack, serial_number=s_n, product_number=p_n, asset_number=a_n, bts_info_id=BTS_INFO.objects.get(bts_name=bts_info), tm500_pc_id=TM500_PC.objects.get(ip=tm_500))
                        count = 1;

                        BTS_PC.objects.filter(ip=server_name).update(bts_mod1="", bts_mod2="", bts_mod3="", tt_mod1="", tt_mod2="", tt_mod3="")
                        for mm in getBTSMOD:
                            if str(mm.bts_info_id.bts_name) == str(bts_info):
                                if count == 1:
                                    #bts_info_id=BTS_INFO.objects.get(bts_name=bts_info)
                                    BTS_PC.objects.filter(ip=server_name).update(bts_mod1=BTS_MODULES.objects.get(i_d=mm.i_d))
                                elif count == 2:
                                    BTS_PC.objects.filter(ip=server_name).update(bts_mod2=BTS_MODULES.objects.get(i_d=mm.i_d))
                                elif count == 3:
                                    BTS_PC.objects.filter(ip=server_name).update(bts_mod3=BTS_MODULES.objects.get(i_d=mm.i_d))
                                count = count + 1

                    elif u_e == 'Test Terminal':
                        t_t = request.POST['editTT']
                        # print("HELLOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOO")
                        BTS_PC.objects.filter(ip=server_name).update(display_name=display_name, username=ussername, password=pass_word, domain=do_main, ue_type=u_e, switch_port=sport, notes=txta, rack=rack, serial_number=s_n, product_number=p_n, asset_number=a_n, bts_info_id=BTS_INFO.objects.get(bts_name=bts_info), tt_info_id=BTS_INFO.objects.get(bts_name=t_t))
                        countbts = 1;
                        counttt = 1;

                        BTS_PC.objects.filter(ip=server_name).update(bts_mod1="", bts_mod2="", bts_mod3="", tt_mod1="", tt_mod2="", tt_mod3="")
                        for mm in getBTSMOD:
                            if str(mm.bts_info_id.bts_name) == str(bts_info):
                                if countbts == 1:
                                    #bts_info_id=BTS_INFO.objects.get(bts_name=bts_info)
                                    BTS_PC.objects.filter(ip=server_name).update(bts_mod1=BTS_MODULES.objects.get(i_d=mm.i_d))
                                elif countbts == 2:
                                    BTS_PC.objects.filter(ip=server_name).update(bts_mod2=BTS_MODULES.objects.get(i_d=mm.i_d))
                                elif countbts == 3:
                                    BTS_PC.objects.filter(ip=server_name).update(bts_mod3=BTS_MODULES.objects.get(i_d=mm.i_d))
                                countbts = countbts + 1
                            if str(mm.bts_info_id.bts_name) == str(t_t):
                                if counttt == 1:
                                    #bts_info_id=BTS_INFO.objects.get(bts_name=bts_info)
                                    BTS_PC.objects.filter(ip=server_name).update(tt_mod1=BTS_MODULES.objects.get(i_d=mm.i_d))
                                elif counttt == 2:
                                    BTS_PC.objects.filter(ip=server_name).update(tt_mod2=BTS_MODULES.objects.get(i_d=mm.i_d))
                                elif counttt == 3:
                                    BTS_PC.objects.filter(ip=server_name).update(tt_mod3=BTS_MODULES.objects.get(i_d=mm.i_d))
                                counttt = counttt + 1
                    messages.success(request, 'Successfully Updated')
                except IntegrityError as e:
                    messages.error(request, 'ERROR')
            except:
                try:
                    if u_e == 'TM500':
                        tm_500 = request.POST['editTM500']
                        BTS_PC.objects.filter(ip=server_name).update(display_name=display_name, username=ussername, password=pass_word, domain=do_main, ue_type=u_e, switch_port=sport, notes=txta, rack=rack, serial_number=s_n, product_number=p_n, asset_number=a_n, tm500_pc_id=TM500_PC.objects.get(ip=tm_500))
                    elif u_e == 'Test Terminal':
                        t_t = request.POST['editTT']
                        BTS_PC.objects.filter(ip=server_name).update(display_name=display_name, username=ussername, password=pass_word, domain=do_main, ue_type=u_e, switch_port=sport, notes=txta, rack=rack, serial_number=s_n, product_number=p_n, asset_number=a_n, tt_info_id=BTS_INFO.objects.get(bts_name=t_t))
                        countbts = 1;
                        counttt = 1;

                        BTS_PC.objects.filter(ip=server_name).update(bts_mod1="", bts_mod2="", bts_mod3="", tt_mod1="", tt_mod2="", tt_mod3="")
                        for mm in getBTSMOD:
                            if str(mm.bts_info_id.bts_name) == str(t_t):
                                if counttt == 1:
                                    #bts_info_id=BTS_INFO.objects.get(bts_name=bts_info)
                                    BTS_PC.objects.filter(ip=server_name).update(tt_mod1=BTS_MODULES.objects.get(i_d=mm.i_d))
                                elif counttt == 2:
                                    BTS_PC.objects.filter(ip=server_name).update(tt_mod2=BTS_MODULES.objects.get(i_d=mm.i_d))
                                elif counttt == 3:
                                    BTS_PC.objects.filter(ip=server_name).update(tt_mod3=BTS_MODULES.objects.get(i_d=mm.i_d))
                                counttt = counttt + 1
                    messages.success(request, 'Successfully Updated')
                except IntegrityError as e:
                    messages.error(request, 'ERROR')

        elif operation == "del-indiv-bts":
            dip = request.POST['deleteID']
            try:
                if BTS_PC.objects.filter(ip = dip).exists():
                    BTS_PC.objects.get(ip = dip).delete()
                messages.success(request, 'Successfully Deleted')
            except:
                  messages.error(request, 'Cannot Delete BTS PC')         

        elif operation == "delete-bts":
            server_list = request.POST.getlist('server_name')
            # print("LIST: ", len(server_list))
            if(len(server_list) < 1):
                messages.error(request, 'Nothing to Delete - Nothing Selected')
            else:
                try:
                    for ip in server_list:
                        print(ip)
                        if BTS_PC.objects.filter(ip = ip).exists():
                            BTS_PC.objects.get(ip = ip).delete()
                    messages.success(request, 'Successfully Deleted')
                except:
                    messages.error(request, 'Cannot Delete BTS PCs')

        ##############################################################
        ################### BTS INFO OPERATIONS ######################
        ##############################################################

        elif operation == "add-info":
            name = request.POST['nam_e']
            typ_e = request.POST['type']
            use = request.POST['use']
            rack = request.POST['rack']
            sport = request.POST['sport']
            txta = request.POST['txta']
            inmod = request.POST.getlist('inmod')

            #print(name, typ_e, use, rack, sport, txta)
            #arr = []
            #index = 0;
            '''for inf in getBTSINFO:
                #print(inf.i_d)
                curr = inf.i_d
                arr.append(curr)
            for i in range(len(arr)):
                # print("arr[i]+1: ", arr[i]+1)
                #print("arr[i+1]: ", arr[i+1])
                #print(i, len(arr))
                if i == len(arr)-1:
                    #print("equal")s
                    index = len(arr)+1
                    arr.append(len(arr)+1)
                    #print("up: ", index)
                    break;
                else:
                    if arr[i]+1 != arr[i+1]:
                        index = arr[i]+1
                        arr.append(arr[i]+1)
                        #print("down: ", index)
                        break
                #arr.sort()
                #print(arr)'''
            try:
                # print(index)
                newInfo = BTS_INFO(bts_name=name, bts_type=typ_e, bts_use=use, rack=rack, switch_port=sport, notes=txta)
                newInfo.save()
                print(inmod)
                for i in range(len(inmod)):
                    print("mod: ", inmod[i])
                    nmod = inmod[i]
                    #field_val = getattr(BTS_MODULES.objects.get(i_d=i), 'bts_name')
                    #print(field_val)     
                    print("module: ", BTS_MODULES.objects.get(i_d=nmod))    
                    print("info: ", BTS_INFO.objects.get(bts_name=name))       
                    BTS_MODULES.objects.filter(i_d=nmod).update(assign=1, bts_info_id=BTS_INFO.objects.get(bts_name=name))
                messages.success(request, 'Successfully Added')
            except:
                messages.error(request, 'Cannot Add BTS Info')

        elif operation == "update-info":
            server_name = request.POST['editID']
            name = request.POST['editName']
            typ_e = request.POST['editType']
            use = request.POST['editUse']
            rack = request.POST['editRack']
            sport = request.POST['editSPort']
            txta = request.POST['editNotes']
            inmod = request.POST.getlist('editMod')

            try:
                # newInfo = BTS_INFO(i_d=index, bts_name=name, bts_type=typ_e, bts_use=use, rack=rack, switch_port=sport, notes=txta)
                BTS_INFO.objects.filter(i_d=server_name).update(bts_name=name, bts_type=typ_e, bts_use=use, rack=rack, switch_port=sport, notes=txta)
                counter = 0
                for i in range(len(inmod)):
                    #print(i, len(inmod))
                    #print("mod: ", inmod[i])
                    # print(len(inmod))
                    nmod = inmod[i]
                    for mm in getBTSMOD:
                        #print(mm)
                        if str(mm.i_d) == str(nmod):
                            if mm.assign == '0':
                                BTS_MODULES.objects.filter(i_d=nmod).update(assign=1, bts_info_id=BTS_INFO.objects.get(bts_name=name))
                                #print("direct save")
                                break
                            elif mm.assign == '1':
                                counter = counter + 1
                                #print("not allowed")
                                break
                #print("##################################################################")
                for nn in getBTSMOD:
                    #print(nn)
                    if str(nn.bts_info_id) == str(name):
                        #print(nn.bts_info_id, " == ", name)
                        #print("module: ", nn.module_name, "(", nn.i_d, ")")
                        for ii in range(len(inmod)):
                            mmod = inmod[ii]
                            #print(mmod)
                            #print(ii, len(inmod))
                            if ii == (len(inmod)-1):
                                #print("lor: ", nn.i_d, mmod)
                                if str(nn.i_d) == str(mmod):
                                    #print("com: ", nn.i_d, mmod)
                                    #print("let it go~")
                                    break
                                else:
                                    #print("end")
                                    BTS_MODULES.objects.filter(i_d=nn.i_d).update(assign=0, bts_info_id="")
                                    break
                            else:
                                if str(nn.i_d) == str(mmod):
                                    #print("com: ", nn.i_d, mmod)
                                    #print("let it go~")
                                    break
                    '''for ii in range(len(inmod)):
                        mmod = inmod[i]
                        
                            if str(nn.i_d) == str(mmod):
                                print("let it go")
                                break
                    print("delete ", nn.)'''
                messages.success(request, 'Successfully Updated')
            except:
                messages.error(request, 'ERROR')
        
        elif operation == "del-indiv-info":
            dip = request.POST['deleteID']
            print(dip)
            try:
                if BTS_INFO.objects.filter(i_d = dip).exists():
                    BTS_INFO.objects.get(i_d = dip).delete()
                messages.success(request, 'Successfully Deleted')
            except:
                messages.error(request, 'Cannot Delete BTS Info')

        elif operation == "delete-info":
            server_list = request.POST.getlist('server_name')
            if(len(server_list) < 1):
                messages.error(request, 'Nothing to Delete - Nothing Selected')
            else:
                try:
                    for ip in server_list:
                        print(ip)
                        if BTS_INFO.objects.filter(i_d = ip).exists():
                            print("delete?")
                            BTS_INFO.objects.get(i_d = ip).delete()
                    messages.success(request, 'Successfully Deleted')
                except:
                    messages.error(request, 'Cannot Delete BTS Info')

        ##############################################################
        ################### BTS MOD OPERATIONS ######################
        ##############################################################

        elif operation == "add-mod":
            # print("oki doki")
            name = request.POST['nam_e']
            b_rand = request.POST['b_rand']
            s_n = request.POST['s_n']
            p_n = request.POST['p_n']
            a_n = request.POST['a_n']
            txta = request.POST['txta']

            #print(name, b_rand, ayd, s_n, p_n, a_n, txta)
            try:
                try:
                    ayd = request.POST['ayd']
                    # print(BTS_INFO.objects.get(i_d=ayd))
                    newMod = BTS_MODULES(module_name=name, module_brand=b_rand, bts_info_id=BTS_INFO.objects.get(i_d=ayd), serial_number=s_n, product_number=p_n, asset_number=a_n, notes=txta, assign=1)
                    newMod.save()
                    messages.success(request, 'Successfully Added')
                except:
                    newMod = BTS_MODULES(module_name=name, module_brand=b_rand, serial_number=s_n, product_number=p_n, asset_number=a_n, notes=txta, assign=0)
                    newMod.save()
                    messages.success(request, 'Successfully Added')
            except:
                messages.error(request, 'ERROR')

        elif operation == "update-mod":
            # print("oki doki")
            i_d = request.POST['editID']
            name = request.POST['editName']
            b_rand = request.POST['editBrand']
            s_n = request.POST['editSNum']
            p_n = request.POST['editPNum']
            a_n = request.POST['editANum']
            txta = request.POST['editNotes']

            try:
                try:
                    ayd = request.POST['editBITIN']
                    print(ayd)
                    BTS_MODULES.objects.filter(i_d=i_d).update(module_name=name, module_brand=b_rand, bts_info_id=BTS_INFO.objects.get(bts_name=ayd), serial_number=s_n, product_number=p_n, asset_number=a_n, notes=txta, assign=1)
                    #newMod.save()
                    messages.success(request, 'Successfully Updated')
                except:
                    #messages.error(request, 'ERROR')
                    BTS_MODULES.objects.filter(i_d=i_d).update(module_name=name, module_brand=b_rand, serial_number=s_n, product_number=p_n, asset_number=a_n, notes=txta, assign=0)
                    newMod.save()
                    messages.success(request, 'Successfully Added')
            except:
                messages.error(request, 'ERROR')

        elif operation == "del-indiv-mod":
            dip = request.POST['deleteID']
            print(dip)
            try:
                if BTS_MODULES.objects.filter(i_d = dip).exists():
                    #BTS_MODULES.objects.get(i_d = dip).delete()
                    field_val = getattr(BTS_MODULES.objects.get(i_d=dip), 'assign')
                    print(field_val)
                    if field_val == "0":
                        BTS_MODULES.objects.get(i_d = dip).delete()
                        messages.success(request, 'Successfully Deleted')
                    else:
                        messages.error(request, 'Cannot Delete BTS Module', extra_tags='error_delete_module')
            except:
                messages.error(request, 'Cannot Delete BTS Module')
        
        elif operation == "delete-mod":
            server_list = request.POST.getlist('server_name')
            if(len(server_list) < 1):
                messages.error(request, 'Nothing to Delete - Nothing Selected')
            else:
                try: 
                    for ip in server_list:
                        print(ip)
                        if BTS_MODULES.objects.filter(i_d = ip).exists():
                            print("delete?")
                            BTS_MODULES.objects.get(i_d = ip).delete()
                    messages.success(request, 'Successfully Deleted')
                except:
                    messages.error(request, 'Cannot Delete BTS Module')

        ##############################################################
        ################### TM500 PC OPERATIONS ######################
        ##############################################################

        elif operation == "add-tm500":
            server_name = request.POST['server_name']
            display_name = request.POST['display_name']
            ussername = request.POST['username']
            pass_word = request.POST['password']
            do_main = request.POST['do_main']
            sport = request.POST['sport']
            txta = request.POST['txta']
            rack = request.POST['rack']
            s_n = request.POST['s_n']
            p_n = request.POST['p_n']
            a_n = request.POST['a_n']

            print(server_name, display_name, ussername, pass_word, do_main, sport, txta, rack)
            try:
                tm_info_ip = request.POST['tm_info_ip']
                print(tm_info_ip)
                try:
                    #newTM = TM500_PC(ip=server_name)
                    newTM = TM500_PC(ip=server_name, display_name=display_name, username=ussername, password=ussername, domain=do_main, tm500_info_id=TM500_INFO.objects.get(ip=tm_info_ip), rack=txta, switch_port=sport, notes=txta, serial_number=s_n, product_number=p_n, asset_number=a_n)
                    newTM.save()
                    messages.success(request, 'Successfully Added')
                except:
                    messages.error(request, 'Cannot Add TM500 PC', extra_tags="tm500_add_error")
            except:
                #print("okay")
                try:
                    #newTM = TM500_PC(ip=server_name)
                    newTM = TM500_PC(ip=server_name, display_name=display_name, username=ussername, password=ussername, domain=do_main, rack=txta, switch_port=sport, notes=txta, serial_number=s_n, product_number=p_n, asset_number=a_n)
                    newTM.save()
                    messages.success(request, 'Successfully Added')
                except:
                    messages.error(request, 'Cannot Add TM500 PC', extra_tags="tm500_add_error")

        elif operation == "update-tm500":
            server_name = request.POST['editID']
            display_name = request.POST['editDN']
            ussername = request.POST['editUN']
            pass_word = request.POST['editPW']
            do_main = request.POST['editDM']
            sport = request.POST['editSPort']
            txta = request.POST['editNotes']
            rack = request.POST['editRack']
            s_n = request.POST['editSNum']
            p_n = request.POST['editPNum']
            a_n = request.POST['editANum']

            try:
                tm_info_ip = request.POST['tm_info_ip']
                #print(tm_info_ip)
                try:
                    TM500_PC.objects.filter(ip=server_name).update(display_name=display_name, username=ussername, password=ussername, domain=do_main, tm500_info_id=TM500_INFO.objects.get(ip=tm_info_ip), rack=txta, switch_port=sport, notes=txta, serial_number=s_n, product_number=p_n, asset_number=a_n)
                    messages.success(request, 'Successfully Updated')
                except:
                    messages.error(request, 'ERROR')
            except:
                try: 
                    TM500_PC.objects.filter(ip=server_name).update(display_name=display_name, username=ussername, password=ussername, domain=do_main, rack=txta, switch_port=sport, notes=txta, serial_number=s_n, product_number=p_n, asset_number=a_n)
                    messages.success(request, 'Successfully Updated')
                except:
                    messages.error(request, 'ERROR')

        elif operation == "del-indiv-tmpc":
            dip = request.POST['deleteID']
            print(dip)
            try:
                if TM500_PC.objects.filter(ip = dip).exists():
                    TM500_PC.objects.get(ip = dip).delete()
                    messages.success(request, 'Successfully Deleted')
            except:
                messages.error(request, 'Cannot Delete TM500 PC', extra_tags="tm500_del_error")

        elif operation == "delete-tmpc":
            server_list = request.POST.getlist('server_name')
            if(len(server_list) < 1):
                messages.error(request, 'Nothing to Delete - Nothing Selected')
            else:
                try: 
                    for ip in server_list:
                        print(ip)
                        if TM500_PC.objects.filter(ip = ip).exists():
                            print("delete?")
                            TM500_PC.objects.get(ip = ip).delete()
                    messages.success(request, 'Successfully Deleted')
                except:
                    messages.error(request, 'Cannot Delete TM500 PCs')

        ##############################################################
        ################### TM500 INFO OPERATIONS ####################
        ##############################################################

        elif operation == "add-tm-info":
            server_name = request.POST['server_name']
            sport = request.POST['sport']
            txta = request.POST['txta']
            rack = request.POST['rack']
            s_n = request.POST['s_n']
            p_n = request.POST['p_n']
            a_n = request.POST['a_n']

            try:
                newTF = TM500_INFO(ip=server_name, switch_port=sport, notes=txta, rack=rack, serial_number=s_n, product_number=p_n, asset_number=a_n)
                newTF.save()
                messages.success(request, "Successfully Added")
            except:
                messages.error(request, "Cannot ADD TM500 INFO")

        elif operation == "update-tm-info":
            server_name = request.POST['editID']
            sport = request.POST['editSPort']
            txta = request.POST['editNotes']
            rack = request.POST['editRack']
            s_n = request.POST['editSNum']
            p_n = request.POST['editPNum']
            a_n = request.POST['editANum']

            try:
                TM500_INFO.objects.filter(ip=server_name).update(switch_port=sport, notes=txta, rack=rack, serial_number=s_n, product_number=p_n, asset_number=a_n)
                messages.success(request, "Successfully Updated")
            except:
                messages.error(request, "ERROR")

        elif operation == "del-indiv-tm-info":
            dip = request.POST['deleteID']
            print(dip)
            try:
                if TM500_INFO.objects.filter(ip = dip).exists():
                    TM500_INFO.objects.get(ip = dip).delete()
                    messages.success(request, 'Successfully Deleted')
            except:
                messages.error(request, 'Cannot Delete TM500 INFO', extra_tags="tminfo_del_error")

        elif operation == "delete-tminfo":
            server_list = request.POST.getlist('server_name')
            if(len(server_list) < 1):
                messages.error(request, 'Nothing to Delete - Nothing Selected')
            else:
                try: 
                    for ip in server_list:
                        print(ip)
                        if TM500_INFO.objects.filter(ip = ip).exists():
                            print("delete?")
                            TM500_INFO.objects.get(ip = ip).delete()
                    messages.success(request, 'Successfully Deleted')
                except:
                    messages.error(request, 'Cannot Delete TM500 INFOs')


    context = {
        "getUE" : getUE,
        "getBTSPC" : getBTSPC,
        "getTMPC" : getTMPC,
        "getBTSINFO" : getBTSINFO,
        "getBTSMOD" : getBTSMOD,
        "getTMINFO" : getTMINFO,
    }
    return render(request, 'index.html', context)

def bts(request):
    getBTS = btsList.objects.all()

    '''if request.method == "POST":
        operation = request.POST['operation']

        # add new bts
        if operation == "add":
            new_bts = request.POST['b_t_s']
            if btsList.objects.filter(bts = new_bts).exists():
                messages.error(request, "Cannot add new BTS")
            else:
                try:
                    newBTS = btsList(bts = new_bts)
                    newBTS.save()
                except IntegrityError as e:
                    messages.error(request, "Cannot add new BTS")

        # edit bts
        elif operation == "update":
            new_bts = request.POST['editBTS']
            old_bts = request.POST['oldBTS']
            # print(old_bts)
            btsList.objects.filter(bts = old_bts).update(bts = new_bts)
            messages.success(request, 'Successfully Updated')

        # delete single
        elif operation == "del_indiv":
            dbts = request.POST['deleteBTS']
            btsList.objects.get(bts = dbts).delete()

        # delete multiple
        elif operation == "delete":
            bts_list = request.POST.getlist('b_t_s')
            for ip in bts_list:
                # print(ip)
                btsList.objects.get(bts = ip).delete()'''
    context = {
        "getBTS" : getBTS,
    }
    return render(request, 'bts.html', context)

def ue(request):
    getUE = ueList.objects.all()

    if request.method == "POST":
        operation = request.POST['operation']

        # add new ue
        if operation == "add":
            new_ue = request.POST['u_e']
            if ueList.objects.filter(ue = new_ue).exists():
                messages.error(request, "Cannot add new UE")
            else:
                try:
                    newUE = ueList(ue = new_ue)
                    newUE.save()
                except IntegrityError as e:
                    messages.error(request, "Cannot add new UE")

        # edit ue info
        elif operation == "update":
            new_ue = request.POST['editUE']
            old_ue = request.POST['oldUE']
            # print(old_bts)
            ueList.objects.filter(ue = old_ue).update(ue = new_ue)
            messages.success(request, 'Successfully Updated')

        # delete single
        elif operation == "del_indiv":
            due = request.POST['deleteUE']
            ueList.objects.get(ue = due).delete()

        # delete multiple
        elif operation == "delete":
            ue_list = request.POST.getlist('u_e')
            for ip in ue_list:
                # print(ip)
                ueList.objects.get(ue = ip).delete()
    context = {
        "getUE" : getUE,
    }
    return render(request, 'ue.html', context)

def racks(request):
    return render(request, 'racks.html')

def exportCSV_BTSPC(request):
    category_queryset = BTS_PC.objects.all()
    response = HttpResponse(
        content_type='application/ms-excel',
    )
    response['Content-Disposition'] = 'attachment; filename="bts-pc.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    row = 1
    col = 0

    ws.cell(row=row, column=1).value = "IP"
    ws.cell(row=row, column=2).value = "Display Name"
    ws.cell(row=row, column=3).value = "Username"
    ws.cell(row=row, column=4).value = "Password"
    ws.cell(row=row, column=5).value = "Domain"
    ws.cell(row=row, column=6).value = "BTS Info"
    ws.cell(row=row, column=7).value = "UE Type"
    ws.cell(row=row, column=8).value = "Rack"
    ws.cell(row=row, column=9).value = "Test Terminal Info"
    ws.cell(row=row, column=10).value = "TM500 Info"
    ws.cell(row=row, column=11).value = "Switch Port"
    ws.cell(row=row, column=12).value = "Serial No."
    ws.cell(row=row, column=13).value = "Product No."
    ws.cell(row=row, column=14).value = "Asset No."
    ws.cell(row=row, column=15).value = "Notes"

    row = 2
    for rpc in category_queryset:
        ws.cell(row=row, column=1).value = rpc.ip
        ws.cell(row=row, column=2).value = rpc.display_name
        ws.cell(row=row, column=3).value = rpc.username
        ws.cell(row=row, column=4).value = rpc.password
        ws.cell(row=row, column=5).value = rpc.domain
        # bts
        if rpc.bts_info_id is None:
             ws.cell(row=row, column=6).value = ""
        else:
            ws.cell(row=row, column=6).value = rpc.bts_info_id.bts_name
        ws.cell(row=row, column=7).value = rpc.ue_type
        ws.cell(row=row, column=8).value = rpc.rack
        # test terminal
        if rpc.tt_info_id is None:
             ws.cell(row=row, column=9).value = ""
        else:
            ws.cell(row=row, column=9).value = rpc.tt_info_id.bts_name
        # tm500
        if rpc.tm500_pc_id is None:
             ws.cell(row=row, column=10).value = ""
        else:
            ws.cell(row=row, column=10).value = rpc.tm500_pc_id.ip
        ws.cell(row=row, column=11).value = rpc.switch_port
        ws.cell(row=row, column=12).value = rpc.serial_number
        ws.cell(row=row, column=13).value = rpc.product_number
        ws.cell(row=row, column=14).value = rpc.asset_number
        ws.cell(row=row, column=15).value = rpc.notes
        row = row + 1

    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font

    wb.save(response)
    return response

def exportCSV_BTSINFO(request):
    category_queryset = BTS_INFO.objects.all()
    response = HttpResponse(
        content_type='application/ms-excel',
    )
    response['Content-Disposition'] = 'attachment; filename="bts-info.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    row = 1
    col = 0

    ws.cell(row=row, column=1).value = "ID"
    ws.cell(row=row, column=2).value = "Name"
    ws.cell(row=row, column=3).value = "Type"
    ws.cell(row=row, column=4).value = "Use"
    ws.cell(row=row, column=5).value = "Rack"
    ws.cell(row=row, column=6).value = "Switch Port"
    ws.cell(row=row, column=7).value = "Modules"
    ws.cell(row=row, column=8).value = "Notes"

    row = 2
    for rpc in category_queryset:
        ws.cell(row=row, column=1).value = rpc.i_d
        ws.cell(row=row, column=2).value = rpc.bts_name
        ws.cell(row=row, column=3).value = rpc.bts_type
        ws.cell(row=row, column=4).value = rpc.bts_use
        ws.cell(row=row, column=5).value = rpc.rack
        ws.cell(row=row, column=6).value = rpc.switch_port

        mod_queryset = BTS_MODULES.objects.all()
        st = ""
        for mm in mod_queryset:
            if mm.bts_info_id.i_d == rpc.i_d:
                st = st + str(mm.module_name)
                st = st + "; "

        ws.cell(row=row, column=7).value = st
        ws.cell(row=row, column=8).value = rpc.notes
        
        row = row + 1

    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font

    wb.save(response)
    return response

def exportCSV_BTSMOD(request):
    category_queryset = BTS_MODULES.objects.all()
    response = HttpResponse(
        content_type='application/ms-excel',
    )
    response['Content-Disposition'] = 'attachment; filename="bts-modules.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    row = 1
    col = 0

    ws.cell(row=row, column=1).value = "ID"
    ws.cell(row=row, column=2).value = "BTS Info ID"
    ws.cell(row=row, column=3).value = "Name"
    ws.cell(row=row, column=4).value = "Brand"
    ws.cell(row=row, column=5).value = "Serial No."
    ws.cell(row=row, column=6).value = "Product No."
    ws.cell(row=row, column=7).value = "Asset No."
    ws.cell(row=row, column=8).value = "Notes"

    row = 2
    for rpc in category_queryset:
        ws.cell(row=row, column=1).value = rpc.i_d

        if rpc.bts_info_id is None:
            ws.cell(row=row, column=2).value = ""
        else:
            ws.cell(row=row, column=2).value = rpc.bts_info_id.bts_name

        ws.cell(row=row, column=3).value = rpc.module_name
        ws.cell(row=row, column=4).value = rpc.module_brand
        ws.cell(row=row, column=5).value = rpc.serial_number
        ws.cell(row=row, column=6).value = rpc.product_number
        ws.cell(row=row, column=7).value = rpc.asset_number
        ws.cell(row=row, column=8).value = rpc.notes

        row = row + 1

    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font

    wb.save(response)
    return response

def exportCSV_TMPC(request):
    category_queryset = TM500_PC.objects.all()
    response = HttpResponse(
        content_type='application/ms-excel',
    )
    response['Content-Disposition'] = 'attachment; filename="tm500-pc.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    row = 1
    col = 0

    ws.cell(row=row, column=1).value = "IP"
    ws.cell(row=row, column=2).value = "Display Name"
    ws.cell(row=row, column=3).value = "Username"
    ws.cell(row=row, column=4).value = "Password"
    ws.cell(row=row, column=5).value = "Domain"
    ws.cell(row=row, column=6).value = "TM500 Info IP"
    ws.cell(row=row, column=7).value = "Rack"
    ws.cell(row=row, column=8).value = "Switch Port"
    ws.cell(row=row, column=9).value = "Serial No."
    ws.cell(row=row, column=10).value = "Product No."
    ws.cell(row=row, column=11).value = "Asset No."
    ws.cell(row=row, column=12).value = "Notes"

    row = 2
    for rpc in category_queryset:
        ws.cell(row=row, column=1).value = rpc.ip
        ws.cell(row=row, column=2).value = rpc.display_name
        ws.cell(row=row, column=3).value = rpc.username
        ws.cell(row=row, column=4).value = rpc.password
        ws.cell(row=row, column=5).value = rpc.domain

        if rpc.tm500_info_id is None:
            ws.cell(row=row, column=6).value = ""
        else:
            ws.cell(row=row, column=6).value = rpc.tm500_info_id.ip
        
        ws.cell(row=row, column=7).value = rpc.rack
        ws.cell(row=row, column=8).value = rpc.switch_port
        ws.cell(row=row, column=9).value = rpc.serial_number
        ws.cell(row=row, column=10).value = rpc.product_number
        ws.cell(row=row, column=11).value = rpc.asset_number
        ws.cell(row=row, column=12).value = rpc.notes

        row = row + 1

    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font

    wb.save(response)
    return response

def exportCSV_TMINFO(request):
    category_queryset = TM500_INFO.objects.all()
    response = HttpResponse(
        content_type='application/ms-excel',
    )
    response['Content-Disposition'] = 'attachment; filename="tm500-info.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    row = 1
    col = 0

    ws.cell(row=row, column=1).value = "IP"
    ws.cell(row=row, column=2).value = "Rack"
    ws.cell(row=row, column=3).value = "Switch Port"
    ws.cell(row=row, column=4).value = "Serial No."
    ws.cell(row=row, column=5).value = "Product No."
    ws.cell(row=row, column=6).value = "Asset No."
    ws.cell(row=row, column=7).value = "Notes"

    row = 2
    for rpc in category_queryset:
        ws.cell(row=row, column=1).value = rpc.ip
        ws.cell(row=row, column=2).value = rpc.rack
        ws.cell(row=row, column=3).value = rpc.switch_port
        ws.cell(row=row, column=4).value = rpc.serial_number
        ws.cell(row=row, column=5).value = rpc.product_number
        ws.cell(row=row, column=6).value = rpc.asset_number
        ws.cell(row=row, column=7).value = rpc.notes

        row = row + 1

    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font

    wb.save(response)
    return response